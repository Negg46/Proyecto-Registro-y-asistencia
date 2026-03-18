from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from Conexion import Registro_datos
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

app = Flask(__name__)
CORS(app)

db = Registro_datos()

@app.route("/validar", methods=["POST"])
def validar():
    try:
        data = request.get_json()
        dni = data.get("identificacion")
        
        if not dni:
            return jsonify({"status": "error", "message": "DNI no proporcionado"}), 400

        resultado = db.registrar_asistencia(dni)
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/")
def inicio():
    return send_from_directory(app.root_path, "index.html")


@app.route("/admin")
def admin():
    return send_from_directory(app.root_path, "dashboard.html")


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


def _serializar_fecha(valor):
    return valor.strftime("%Y-%m-%d") if hasattr(valor, "strftime") else valor


def _serializar_hora(valor):
    if valor is None:
        return None
    if isinstance(valor, timedelta):
        total_segundos = int(valor.total_seconds())
        horas = total_segundos // 3600
        minutos = (total_segundos % 3600) // 60
        segundos = total_segundos % 60
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
    return valor.strftime("%H:%M:%S") if hasattr(valor, "strftime") else valor


def _obtener_payload_dashboard(cliente_id=None):
    estadisticas = db.obtener_estadisticas_dashboard()
    historico = db.obtener_asistencias_ultimos_dias(7)
    top_clientes = db.obtener_top_clientes(5)
    clientes_hoy = db.obtener_clientes_hoy(8)

    payload = {
        "estadisticas": estadisticas,
        "historico": [
            {
                "fecha": _serializar_fecha(fila["fecha"]),
                "total_visitas": fila["total_visitas"],
                "clientes_unicos": fila["clientes_unicos"],
            }
            for fila in historico
        ],
        "top_clientes": top_clientes,
        "clientes_hoy": [
            {
                **fila,
                "ultima_entrada": _serializar_hora(fila["ultima_entrada"]),
            }
            for fila in clientes_hoy
        ],
    }

    if cliente_id:
        cliente = db.obtener_resumen_cliente(cliente_id)
        if cliente:
            payload["cliente"] = {
                **cliente,
                "ultima_hora_hoy": _serializar_hora(cliente["ultima_hora_hoy"]),
            }
            payload["cliente_asistencias"] = [
                {
                    "fecha": _serializar_fecha(fila["fecha"]),
                    "visitas": fila["visitas"],
                }
                for fila in db.obtener_asistencias_por_dia_cliente(cliente_id)
            ]
            payload["cliente_movimientos_hoy"] = [
                {"hora": _serializar_hora(fila["hora"])}
                for fila in db.obtener_movimientos_hoy_cliente(cliente_id)
            ]

    return payload


def _crear_excel_dashboard(payload):
    workbook = Workbook()
    hoja_resumen = workbook.active
    hoja_resumen.title = "Resumen"
    hoja_resumen.append(["Metrica", "Valor"])
    hoja_resumen.append(["Clientes registrados", payload["estadisticas"]["total_clientes"]])
    hoja_resumen.append(["Visitas hoy", payload["estadisticas"]["visitas_hoy"]])
    hoja_resumen.append(["Clientes unicos hoy", payload["estadisticas"]["clientes_unicos_hoy"]])
    hoja_resumen.append(["Clientes VIP", payload["estadisticas"]["total_vip"]])
    hoja_resumen.append(["Clientes Clasica", payload["estadisticas"]["total_clasica"]])

    hoja_historico = workbook.create_sheet("Historico")
    hoja_historico.append(["Fecha", "Visitas", "Clientes unicos"])
    for fila in payload["historico"]:
        hoja_historico.append([fila["fecha"], fila["total_visitas"], fila["clientes_unicos"]])

    hoja_top = workbook.create_sheet("Top clientes")
    hoja_top.append(["Nombre", "Nivel", "Total registros", "Total dias"])
    for fila in payload["top_clientes"]:
        hoja_top.append([fila["nombre"], fila["nivel"], fila["total_registros"], fila["total_dias"]])

    hoja_hoy = workbook.create_sheet("Entradas hoy")
    hoja_hoy.append(["Nombre", "Nivel", "Visitas hoy", "Ultima entrada"])
    for fila in payload["clientes_hoy"]:
        hoja_hoy.append([fila["nombre"], fila["nivel"], fila["visitas_hoy"], fila["ultima_entrada"]])

    cliente = payload.get("cliente")
    if cliente:
        hoja_cliente = workbook.create_sheet("Cliente")
        hoja_cliente.append(["Dato", "Valor"])
        hoja_cliente.append(["Nombre", cliente["nombre"]])
        hoja_cliente.append(["DNI", cliente["dni"]])
        hoja_cliente.append(["Nivel", cliente["nivel"]])
        hoja_cliente.append(["Tarjeta", cliente["tarjeta"]])
        hoja_cliente.append(["Dias asistidos", cliente["total_dias"]])
        hoja_cliente.append(["Registros totales", cliente["total_registros"]])
        hoja_cliente.append(["Visitas hoy", cliente["visitas_hoy"] or 0])

        hoja_cliente_hist = workbook.create_sheet("Cliente historico")
        hoja_cliente_hist.append(["Fecha", "Visitas"])
        for fila in payload.get("cliente_asistencias", []):
            hoja_cliente_hist.append([fila["fecha"], fila["visitas"]])

    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    return salida


def _crear_pdf_dashboard(payload):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    estilos = getSampleStyleSheet()
    contenido = []

    contenido.append(Paragraph("Casino La Rioja - Reporte de Dashboard", estilos["Title"]))
    contenido.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", estilos["Normal"]))
    contenido.append(Spacer(1, 0.4 * cm))

    resumen = [
        ["Metrica", "Valor"],
        ["Clientes registrados", payload["estadisticas"]["total_clientes"]],
        ["Visitas hoy", payload["estadisticas"]["visitas_hoy"]],
        ["Clientes unicos hoy", payload["estadisticas"]["clientes_unicos_hoy"]],
        ["Clientes VIP", payload["estadisticas"]["total_vip"]],
        ["Clientes Clasica", payload["estadisticas"]["total_clasica"]],
    ]
    contenido.append(Paragraph("Resumen general", estilos["Heading2"]))
    contenido.append(_crear_tabla_pdf(resumen))
    contenido.append(Spacer(1, 0.35 * cm))

    historico = [["Fecha", "Visitas", "Clientes unicos"]] + [
        [fila["fecha"], fila["total_visitas"], fila["clientes_unicos"]]
        for fila in payload["historico"]
    ]
    contenido.append(Paragraph("Historico de asistencias", estilos["Heading2"]))
    contenido.append(_crear_tabla_pdf(historico))
    contenido.append(Spacer(1, 0.35 * cm))

    top = [["Nombre", "Nivel", "Registros", "Dias"]] + [
        [fila["nombre"], fila["nivel"], fila["total_registros"], fila["total_dias"]]
        for fila in payload["top_clientes"]
    ]
    contenido.append(Paragraph("Top clientes", estilos["Heading2"]))
    contenido.append(_crear_tabla_pdf(top))

    cliente = payload.get("cliente")
    if cliente:
        contenido.append(Spacer(1, 0.35 * cm))
        contenido.append(Paragraph("Cliente seleccionado", estilos["Heading2"]))
        datos_cliente = [
            ["Dato", "Valor"],
            ["Nombre", cliente["nombre"]],
            ["DNI", cliente["dni"]],
            ["Nivel", cliente["nivel"]],
            ["Tarjeta", cliente["tarjeta"]],
            ["Dias asistidos", cliente["total_dias"]],
            ["Registros totales", cliente["total_registros"]],
            ["Visitas hoy", cliente["visitas_hoy"] or 0],
        ]
        contenido.append(_crear_tabla_pdf(datos_cliente))

        asistencias_cliente = [["Fecha", "Visitas"]] + [
            [fila["fecha"], fila["visitas"]]
            for fila in payload.get("cliente_asistencias", [])
        ]
        contenido.append(Spacer(1, 0.25 * cm))
        contenido.append(_crear_tabla_pdf(asistencias_cliente))

    doc.build(contenido)
    buffer.seek(0)
    return buffer


def _crear_tabla_pdf(datos):
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c91e3a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f7f7f9")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d3d8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#ffffff"), colors.HexColor("#f0f2f5")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tabla


@app.route("/dashboard/resumen", methods=["GET"])
def dashboard_resumen():
    try:
        cliente_id = request.args.get("cliente_id", type=int)
        return jsonify(_obtener_payload_dashboard(cliente_id))
    except Exception as error:
        return jsonify({"status": "error", "message": str(error)}), 500


@app.route("/dashboard/clientes/buscar", methods=["GET"])
def buscar_clientes_dashboard():
    try:
        termino = request.args.get("q", "").strip()
        resultados = db.buscar_clientes(termino)
        return jsonify(resultados)
    except Exception as error:
        return jsonify({"status": "error", "message": str(error)}), 500


@app.route("/dashboard/clientes/<int:cliente_id>", methods=["GET"])
def detalle_cliente_dashboard(cliente_id):
    try:
        cliente = db.obtener_resumen_cliente(cliente_id)
        if not cliente:
            return jsonify({"status": "error", "message": "Cliente no encontrado"}), 404

        asistencias = db.obtener_asistencias_por_dia_cliente(cliente_id)
        movimientos_hoy = db.obtener_movimientos_hoy_cliente(cliente_id)

        return jsonify({
            "cliente": {
                **cliente,
                "ultima_hora_hoy": _serializar_hora(cliente["ultima_hora_hoy"]),
            },
            "asistencias": [
                {
                    "fecha": _serializar_fecha(fila["fecha"]),
                    "visitas": fila["visitas"],
                }
                for fila in asistencias
            ],
            "movimientos_hoy": [
                {"hora": _serializar_hora(fila["hora"])}
                for fila in movimientos_hoy
            ],
        })
    except Exception as error:
        return jsonify({"status": "error", "message": str(error)}), 500


@app.route("/dashboard/export/excel", methods=["GET"])
def exportar_dashboard_excel():
    try:
        cliente_id = request.args.get("cliente_id", type=int)
        salida = _crear_excel_dashboard(_obtener_payload_dashboard(cliente_id))
        return send_file(
            salida,
            as_attachment=True,
            download_name="dashboard_casino_la_rioja.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as error:
        return jsonify({"status": "error", "message": str(error)}), 500


@app.route("/dashboard/export/pdf", methods=["GET"])
def exportar_dashboard_pdf():
    try:
        cliente_id = request.args.get("cliente_id", type=int)
        salida = _crear_pdf_dashboard(_obtener_payload_dashboard(cliente_id))
        return send_file(
            salida,
            as_attachment=True,
            download_name="dashboard_casino_la_rioja.pdf",
            mimetype="application/pdf"
        )
    except Exception as error:
        return jsonify({"status": "error", "message": str(error)}), 500

if __name__ == "__main__":
    import os
    # Captura el puerto que Render asigna, si no existe usa el 10000
    port = int(os.environ.get("PORT", 10000))
    # Importante: host='0.0.0.0' para que sea accesible desde internet
    app.run(host='0.0.0.0', port=port)
